import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
#DeliveryList(2025-12-15)_(0)
# HTML 파일 읽기
with open("스마트로젠.html", "r", encoding="utf-8") as f:
    html = f.read()

soup = BeautifulSoup(html, "html.parser")

# 번호(연락처)를 기준으로 최신 tr만 저장
latest = {}  # "010-xxxx-yyyy" -> {"name": ..., "invoices": ...}

# <tr> 순서대로 읽되 → 동일 번호면 최신값으로 덮어쓰기
for tr in soup.find_all("tr"):
    tds = tr.find_all("td")
    if len(tds) < 3:
        continue

    name = tds[0].get_text(strip=True)
    phone = tds[1].get_text(strip=True)

    invoices = []
    for td in tds[2:]:
        text = td.get_text(strip=True)
        # 송장번호 패턴: 000-0000-0000
        if re.fullmatch(r"\d{3}-\d{4}-\d{4}", text):
            invoices.append(text)

    if not invoices:
        continue

    # 동일 번호면 최신값으로 덮어쓰기
    latest[phone] = {
        "name": name,
        "invoices": invoices
    }

# -----------------------------
#  🔽🔽  여기서부터 엑셀 저장 코드 추가  🔽🔽
# -----------------------------

def save_excel(data_dict, filename="송장번호_결과.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "송장번호"

    # 헤더
    ws.append(["","이름", "송장번호"])

    # 데이터 입력
    for phone, info in data_dict.items():
        name = info["name"]
        invoices = ", ".join(info["invoices"])
        ws.append([name, phone, invoices])

    wb.save(filename)
    print(f"엑셀 저장 완료: {filename}")

# 함수 호출
save_excel(latest)

# "DeliveryList(2025-12-15)_(0).xlsx" 파일 열어서 "구매자" 필드 항목들과 "송장번호_결과.xlsx" 파일의 "이름" 항목들과 필드 비교해서 같으면 "DeliveryList(2025-12-15)_(0).xlsx" "운송장번호" 열에 "송장번호"값 입력해서 저장하기 기능추가해줘.

from openpyxl import load_workbook
from datetime import date

def apply_invoices_to_delivery(
    delivery_file=f"DeliveryList({date.today().strftime('%Y-%m-%d')})_(0).xlsx",
    invoice_file="송장번호_결과.xlsx",
    output_file="DeliveryList_송장입력완료.xlsx"
):
    pass

    # 1️⃣ 송장번호_결과.xlsx 읽기 → 이름:송장번호 dict
    wb_inv = load_workbook(invoice_file)
    ws_inv = wb_inv.active

    invoice_map = {}  # 이름 -> 송장번호
    for row in ws_inv.iter_rows(min_row=2, values_only=True):
        _, name, invoices = row   # ← 여기 중요!!
        if name and invoices:
            invoice_map[name] = invoices
    
    # 2️⃣ DeliveryList 엑셀 열기
    wb_del = load_workbook(delivery_file)
    ws_del = wb_del.active

    # 3️⃣ 헤더 위치 찾기
    header = {cell.value: idx for idx, cell in enumerate(ws_del[1], start=1)}

    buyer_col = header.get("수취인이름")
    invoice_col = header.get("운송장번호")

    if not buyer_col or not invoice_col:
        raise ValueError("구매자 또는 운송장번호 열을 찾을 수 없습니다.")

    # 4️⃣ 구매자 이름 비교 후 운송장번호 입력
    # for row in range(2, ws_del.max_row + 1):
    #     buyer = ws_del.cell(row=row, column=buyer_col).value
    #     if buyer in invoice_map:
    #         ws_del.cell(row=row, column=invoice_col).value = int(invoice_map[buyer])
    
    # 4️⃣ 구매자 이름 비교 후 운송장번호 입력
    for row in range(2, ws_del.max_row + 1):
        buyer = ws_del.cell(row=row, column=buyer_col).value
        if buyer in invoice_map:
            cell = ws_del.cell(row=row, column=invoice_col)
            value = str(invoice_map[buyer]).replace("-", "")
            cell.value = float(value) if value.isdigit() else value


    # 5️⃣ 저장
    wb_del.save(output_file)
    print(f"✅ 운송장번호 입력 완료: {output_file}")

apply_invoices_to_delivery()


print()